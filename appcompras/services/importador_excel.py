import openpyxl
from django.db import transaction
from inventario_app.models import Producto, Proveedor, UnidadMedida


class ImportadorExcel:
    """
    Importador profesional con vista previa y confirmación.
    """

    def __init__(self, archivo, proveedor_id, usuario, confirmar=False):
        self.archivo = archivo
        self.proveedor_id = proveedor_id
        self.usuario = usuario
        self.confirmar = confirmar

        self.errores = []
        self.nuevos = []
        self.actualizados = []

        self.creados = 0
        self.actualizados_count = 0

        self.proveedor = None
        self.wb = None
        self.ws = None

    # -------------------------------------------------------------
    # VALIDACIÓN INICIAL
    # -------------------------------------------------------------
    def validar_proveedor(self):
        try:
            self.proveedor = Proveedor.objects.get(id=self.proveedor_id)
        except Proveedor.DoesNotExist:
            self.errores.append("El proveedor seleccionado no existe.")
            return False
        return True

    # -------------------------------------------------------------
    # CARGAR EXCEL
    # -------------------------------------------------------------
    def cargar_excel(self):
        try:
            self.wb = openpyxl.load_workbook(self.archivo)
            self.ws = self.wb.active
        except Exception as e:
            self.errores.append(f"Error al leer el archivo Excel: {str(e)}")
            return False
        return True

    # -------------------------------------------------------------
    # VALIDAR FILA
    # -------------------------------------------------------------
    def validar_fila(self, fila):
        try:
            codigo = str(fila[0].value).strip() if fila[0].value else None
            nombre = str(fila[1].value).strip() if fila[1].value else None
            precio = float(fila[2].value)
            unidad_abrev = str(fila[3].value).strip() if fila[3].value else None
        except Exception:
            self.errores.append(f"Fila inválida: {fila[0].row}")
            return None

        if not nombre:
            self.errores.append(f"Fila {fila[0].row}: El nombre del producto es obligatorio.")
            return None

        if precio <= 0:
            self.errores.append(f"Fila {fila[0].row}: El precio debe ser mayor que cero.")
            return None

        if not unidad_abrev:
            self.errores.append(f"Fila {fila[0].row}: La unidad de medida es obligatoria.")
            return None

        try:
            unidad = UnidadMedida.objects.get(abreviatura__iexact=unidad_abrev)
        except UnidadMedida.DoesNotExist:
            self.errores.append(
                f"Fila {fila[0].row}: La unidad de medida '{unidad_abrev}' no existe."
            )
            return None

        return {
            "codigo": codigo,
            "nombre": nombre,
            "precio": precio,
            "unidad": unidad,
            "fila": fila[0].row,
        }

    # -------------------------------------------------------------
    # ANALIZAR (VISTA PREVIA)
    # -------------------------------------------------------------
    def analizar(self):
        for fila in self.ws.iter_rows(min_row=2):
            datos = self.validar_fila(fila)
            if not datos:
                continue

            producto = None
            if datos["codigo"]:
                producto = Producto.objects.filter(codigo_barras=datos["codigo"]).first()

            if producto:
                self.actualizados.append({
                    "codigo": datos["codigo"],
                    "nombre_actual": producto.nombre_interno,
                    "nombre_nuevo": datos["nombre"],
                    "precio_actual": producto.precio_compra,
                    "precio_nuevo": datos["precio"],
                    "unidad_actual": producto.unidad_medida.abreviatura,
                    "unidad_nueva": datos["unidad"].abreviatura,
                })
            else:
                self.nuevos.append(datos)

    # -------------------------------------------------------------
    # IMPORTACIÓN REAL
    # -------------------------------------------------------------
    def ejecutar_importacion(self):
        for fila in self.ws.iter_rows(min_row=2):
            datos = self.validar_fila(fila)
            if not datos:
                continue

            producto = None
            if datos["codigo"]:
                producto = Producto.objects.filter(codigo_barras=datos["codigo"]).first()

            if producto:
                producto.nombre_interno = datos["nombre"]
                producto.precio_compra = datos["precio"]
                producto.unidad_medida = datos["unidad"]
                producto.proveedor = self.proveedor
                producto.save()
                self.actualizados_count += 1
            else:
                Producto.objects.create(
                    codigo_barras=datos["codigo"],
                    nombre_interno=datos["nombre"],
                    precio_compra=datos["precio"],
                    unidad_medida=datos["unidad"],
                    proveedor=self.proveedor,
                )
                self.creados += 1

    # -------------------------------------------------------------
    # EJECUCIÓN PRINCIPAL
    # -------------------------------------------------------------
    @transaction.atomic
    def ejecutar(self):
        if not self.validar_proveedor():
            return self.resumen()

        if not self.cargar_excel():
            return self.resumen()

        # Vista previa
        if not self.confirmar:
            self.analizar()
            return {
                "errores": self.errores,
                "nuevos": self.nuevos,
                "actualizados": self.actualizados,
                "mostrar_preview": True,
            }

        # Importación real
        self.ejecutar_importacion()
        return self.resumen()

    # -------------------------------------------------------------
    # RESUMEN FINAL
    # -------------------------------------------------------------
    def resumen(self):
        return {
            "errores": self.errores,
            "creados": self.creados,
            "actualizados": self.actualizados_count,
            "mostrar_preview": False,
        }
