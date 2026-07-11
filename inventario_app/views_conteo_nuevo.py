import json
from decimal import Decimal

from django.http import JsonResponse, HttpResponse
from django.shortcuts import render, get_object_or_404
from django.template.loader import render_to_string
from django.views.decorators.csrf import csrf_exempt
from django.utils import timezone
from django.contrib.auth.decorators import login_required

from inventario_app.models import Producto
from inventario_app.models import ConteoSesion, ConteoLinea
from core.servicios.servicio_stock import ServicioStock

# ============================================================
# FUNCIÓN GLOBAL: Calcular totales de una sesión
# ============================================================
def calcular_totales(sesion):
    lineas = ConteoLinea.objects.filter(sesion=sesion)
    return {
        "importe_contado": sum((l.importe_contado for l in lineas), Decimal("0")),
        "importe_aumenta": sum((l.importe_aumenta for l in lineas), Decimal("0")),
        "importe_disminuye": sum((l.importe_disminuye for l in lineas), Decimal("0")),
    }

# ============================================================
# 1. Página índice del módulo
# ============================================================
def conteo_index(request):
    sesiones = ConteoSesion.objects.all().order_by("-id")
    return render(request, "conteo_nuevo/index.html", {"sesiones": sesiones})

# ============================================================
# 2. Crear nueva sesión (VERSIÓN CORRECTA Y BLINDADA)
# ============================================================
@login_required
def nueva_sesion(request):
    sesion = ConteoSesion.objects.create(
        nombre="Nueva sesión de conteo",
        tipo="conteo",
        estado="abierta",
        usuario=request.user,
    )

    # NO crear línea inicial en la base de datos
    lineas = []

    totales = {"importe_contado": 0, "importe_aumenta": 0, "importe_disminuye": 0}

    return render(
        request,
        "conteo_nuevo/sesion.html",
        {"sesion": sesion, "lineas": lineas, "totales": totales},
    )

# ============================================================
# 3. Cargar una sesión existente
# ============================================================
def sesion_conteo(request, sesion_id):
    sesion = get_object_or_404(ConteoSesion, id=sesion_id)
    lineas = ConteoLinea.objects.filter(sesion=sesion)
    totales = calcular_totales(sesion)

    return render(
        request,
        "conteo_nuevo/sesion.html",
        {"sesion": sesion, "lineas": lineas, "totales": totales},
    )

# ============================================================
# 4. Resumen de la sesión
# ============================================================
def resumen_sesion(request, sesion_id):
    sesion = get_object_or_404(ConteoSesion, id=sesion_id)
    lineas = ConteoLinea.objects.filter(sesion=sesion)

    total_productos = lineas.count()
    total_diferencias = sum((l.diferencia for l in lineas), Decimal("0"))
    valor_contado = sum((l.importe_contado for l in lineas), Decimal("0"))
    valor_teorico = sum((l.importe_teorico for l in lineas), Decimal("0"))

    movimientos = MovimientoStock.objects.filter(
        origen__icontains=f"Conteo/Nueva #{sesion_id}"
    ).order_by('-fecha')

    return render(
        request,
        "conteo_nuevo/resumen.html",
        {
            "sesion": sesion,
            "lineas": lineas,
            "total_productos": total_productos,
            "total_diferencias": total_diferencias,
            "valor_contado": valor_contado,
            "valor_teorico": valor_teorico,
            "movimientos": movimientos,
        },
    )

# ============================================================
# 5. API: Búsqueda universal de producto
# ============================================================
from django.db.models import Q

def api_buscar_producto(request):
    raw = request.GET.get("query", "")
    query = str(raw).strip()

    if not query:
        return JsonResponse({"error": "Código vacío o inválido."})

    producto = Producto.objects.filter(
        Q(codigo_barras=query) |
        Q(codigo_interno=query)
    ).first()

    if producto:
        return _producto_json(producto)

    return JsonResponse({"error": "Producto no encontrado"})


def _producto_json(producto):
    return JsonResponse({
        "id": producto.id,
        "codigo_barras": producto.codigo_barras,
        "codigo_interno": producto.codigo_interno,
        "nombre": producto.nombre_interno,
        "precio_coste": float(producto.precio_compra or 0),
        "stock_teorico": float(producto.stock_actual or 0),
    })
# ============================================================
# 6. API: Agregar línea
# ============================================================
@csrf_exempt
def api_agregar_linea(request):
    data = json.loads(request.body)

    sesion = ConteoSesion.objects.get(id=data["sesion_id"])
    producto = Producto.objects.get(id=data["producto_id"])

    if sesion.estado != "abierta":
        return JsonResponse({"error": "La sesión está cerrada."})

    existente = ConteoLinea.objects.filter(sesion=sesion, producto=producto).first()
    if existente:
        html = render_to_string("conteo_nuevo/componentes/linea.html", {"linea": existente})
        return HttpResponse(html)

    linea = ConteoLinea.objects.create(
        sesion=sesion,
        producto=producto,
        codigo_barras=producto.codigo_barras,
        precio_coste=producto.precio_compra,
        stock_teorico=producto.stock_actual,
        stock_contado=Decimal("0"),
        motivo="",
        ubicacion="",
        usuario_modifica=request.user if request.user.is_authenticated else None,
        fecha_modifica=timezone.now(),
    )

    html = render_to_string("conteo_nuevo/componentes/linea.html", {"linea": linea})
    return HttpResponse(html)

# ============================================================
# 7. API: Actualizar línea
# ============================================================
@csrf_exempt
def api_actualizar_linea(request):
    if request.method != "POST":
        return JsonResponse({"error": "Método no permitido"}, status=405)

    try:
        data = json.loads(request.body)
        linea_id = data.get("linea_id")
        stock_raw = str(data.get("stock_contado", "0")).strip()
        motivo = data.get("motivo", "")

        linea = ConteoLinea.objects.get(id=linea_id)
        sesion = linea.sesion

        if sesion.estado not in ["abierta", "reabierta"]:
            return JsonResponse({"error": "La sesión está cerrada."})

        valor_actual = linea.stock_contado or Decimal("0")

        if stock_raw.startswith("+") or stock_raw.startswith("-"):
            try:
                cambio = Decimal(stock_raw)
                stock_contado = valor_actual + cambio
            except:
                return JsonResponse({"error": "Valor inválido."})
        else:
            try:
                stock_contado = Decimal(stock_raw)
            except:
                return JsonResponse({"error": "Valor inválido."})

        linea.stock_contado = stock_contado
        linea.motivo = motivo
        linea.usuario_modifica = request.user if request.user.is_authenticated else None
        linea.fecha_modifica = timezone.now()
        linea.save()

        return JsonResponse({
            "stock_contado": float(linea.stock_contado),
            "diferencia": float(linea.diferencia),
            "importe_contado": float(linea.importe_contado),
            "importe_teorico": float(linea.importe_teorico),
            "importe_aumenta": float(linea.importe_aumenta),
            "importe_disminuye": float(linea.importe_disminuye),
        })

    except Exception as e:
        return JsonResponse({"error": str(e)}, status=400)

# ============================================================
# 8. API: Eliminar línea
# ============================================================
@csrf_exempt
def api_eliminar_linea(request):
    data = json.loads(request.body)

    linea = ConteoLinea.objects.get(id=data["linea_id"])
    sesion = linea.sesion

    if sesion.estado != "abierta":
        return JsonResponse({"error": "No se pueden eliminar líneas."})

    linea.delete()
    return JsonResponse({"ok": True})

# ============================================================
# 9. Cerrar sesión
# ============================================================
def cerrar_sesion(request, sesion_id):
    sesion = get_object_or_404(ConteoSesion, id=sesion_id)

    sesion.estado = "cerrada"
    sesion.fecha_cierre = timezone.now()
    sesion.usuario_cierre = request.user
    sesion.save()

    lineas = ConteoLinea.objects.filter(sesion=sesion)
    totales = calcular_totales(sesion)

    return render(
        request,
        "conteo_nuevo/sesion.html",
        {
            "sesion": sesion,
            "lineas": lineas,
            "totales": totales,
            "mensaje": "La sesión ha sido cerrada."
        },
    )

# ============================================================
# 10. Reabrir sesión
# ============================================================
def reabrir_sesion(request, sesion_id):
    sesion = get_object_or_404(ConteoSesion, id=sesion_id)

    sesion.estado = "reabierta"
    sesion.fecha_reapertura = timezone.now()
    sesion.usuario_reapertura = request.user
    sesion.save()

    lineas = ConteoLinea.objects.filter(sesion=sesion)
    totales = calcular_totales(sesion)

    return render(
        request,
        "conteo_nuevo/sesion.html",
        {
            "sesion": sesion,
            "lineas": lineas,
            "totales": totales,
            "mensaje": "Sesión reabierta."
        },
    )

# ============================================================
# 11. Aplicar diferencias
# ============================================================
def aplicar_diferencias(request, sesion_id):
    sesion = get_object_or_404(ConteoSesion, id=sesion_id)
    lineas = ConteoLinea.objects.filter(sesion=sesion)

    for linea in lineas:
        producto = linea.producto
        
        ServicioStock.ajustar_stock(
            producto=producto,
            cantidad_final=linea.stock_contado,
            origen=f"Conteo/Nueva #{sesion.id}"
        )

    sesion.generar_asientos_conteo(usuario=request.user)
    sesion.generar_asiento_contable_real(usuario=request.user)

    sesion.estado = "aplicada"
    sesion.fecha_aplicacion = timezone.now()
    sesion.save()

    return resumen_sesion(request, sesion_id)

# ============================================================
# 12. Exportar PDF
# ============================================================
from weasyprint import HTML

def exportar_pdf_conteo(request, sesion_id):
    sesion = get_object_or_404(ConteoSesion, id=sesion_id)
    lineas = ConteoLinea.objects.filter(sesion=sesion)

    html = render_to_string(
        "conteo_nuevo/pdf_conteo.html",
        {
            "sesion": sesion,
            "lineas": lineas,
            "total_productos": lineas.count(),
            "total_diferencias": sum((l.diferencia for l in lineas), Decimal("0")),
            "valor_contado": sum((l.importe_contado for l in lineas), Decimal("0")),
            "valor_teorico": sum((l.importe_teorico for l in lineas), Decimal("0")),
            "usuario": request.user.username,
            "fecha": timezone.now(),
        },
    )
    pdf = HTML(string=html).write_pdf()

    response = HttpResponse(pdf, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="conteo_sesion_{sesion.id}.pdf"'
    return response

# ============================================================
# 13. Exportar Excel
# ============================================================
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.cell.cell import MergedCell

def exportar_excel_conteo(request, sesion_id):
    sesion = ConteoSesion.objects.get(id=sesion_id)
    lineas = ConteoLinea.objects.filter(sesion=sesion)

    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen Conteo"

    bold = Font(bold=True)
    center = Alignment(horizontal="center")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    header_fill = PatternFill(start_color="F0F0F0", fill_type="solid")

    ws["A1"] = f"Resumen de Conteo — Sesión #{sesion.id}"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:L1")

    ws["A2"] = f"Generado por: {request.user.username}"
    ws["A3"] = f"Fecha: {timezone.now()}"

    headers = [
        "Código Barras",
        "Código Interno",
        "Producto",
        "Precio Coste",
        "Stock Teórico",
        "Stock Contado",
        "Importe Teórico",
        "Importe Contado",
        "Diferencia",
        "Importe Aumenta",
        "Importe Disminuye",
        "Motivo",
    ]

    ws.append(headers)

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=4, column=col)
        cell.font = bold
        cell.alignment = center
        cell.border = border
        cell.fill = header_fill

    for linea in lineas:
        diferencia = linea.stock_contado - linea.stock_teorico

        if diferencia > 0:
            aumenta = diferencia * linea.precio_coste
            disminuye = 0
        elif diferencia < 0:
            aumenta = 0
            disminuye = abs(diferencia) * linea.precio_coste
        else:
            aumenta = 0
            disminuye = 0

        ws.append([
            linea.codigo_barras,
            linea.producto.codigo_interno,
            linea.producto.nombre_interno,
            float(linea.precio_coste),
            float(linea.stock_teorico),
            float(linea.stock_contado),
            float(linea.importe_teorico),
            float(linea.importe_contado),
            float(diferencia),
            float(aumenta),
            float(disminuye),
            linea.motivo,
        ])

    for row in ws.iter_rows(min_row=5, max_row=ws.max_row):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal="left")

    for col in ws.columns:
        max_length = 0
        column = None

        for cell in col:
            if isinstance(cell, MergedCell):
                continue

            if column is None:
                column = cell.column_letter

            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass

        if column:
            ws.column_dimensions[column].width = max_length + 2

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="conteo_sesion_{sesion.id}.xlsx"'
    wb.save(response)
    return response

# ============================================================
# 14. Importar Excel
# ============================================================
def importar_excel_conteo(request, sesion_id):
    sesion = get_object_or_404(ConteoSesion, id=sesion_id)

    if request.method == "POST":
        archivo = request.FILES.get("archivo")

        if not archivo:
            return render(
                request,
                "conteo_nuevo/importar_excel.html",
                {"sesion": sesion, "errores": ["No se ha seleccionado ningún archivo."]},
            )

        try:
            wb = load_workbook(archivo)
            ws = wb.active
        except:
            return render(
                request,
                "conteo_nuevo/importar_excel.html",
                {"sesion": sesion, "errores": ["Archivo inválido o corrupto."]},
            )

        errores = []
        for fila in ws.iter_rows(min_row=2, values_only=True):
            codigo_barras, codigo_interno, stock_contado, motivo = fila

            if not codigo_barras and not codigo_interno:
                continue

            producto = None

            if codigo_barras:
                producto = Producto.objects.filter(codigo_barras=str(codigo_barras)).first()

            if not producto and codigo_interno:
                producto = Producto.objects.filter(codigo_interno=str(codigo_interno)).first()

            if not producto:
                errores.append(f"Producto no encontrado: {codigo_barras} / {codigo_interno}")
                continue

            try:
                stock_contado = Decimal(str(stock_contado))
            except:
                errores.append(f"Stock inválido para {producto.nombre_interno}")
                continue

            linea, creada = ConteoLinea.objects.get_or_create(
                sesion=sesion,
                producto=producto,
                defaults={
                    "codigo_barras": producto.codigo_barras,
                    "precio_coste": producto.precio_compra,
                    "stock_teorico": producto.stock_actual,
                },
            )

            linea.stock_contado = stock_contado
            linea.motivo = motivo or ""
            linea.save()

        if errores:
            return render(
                request,
                "conteo_nuevo/importar_excel.html",
                {"sesion": sesion, "errores": errores},
            )

        return sesion_conteo(request, sesion_id)

    return render(
        request,
        "conteo_nuevo/importar_excel.html",
        {"sesion": sesion}
    )
