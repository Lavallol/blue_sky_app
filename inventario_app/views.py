from decimal import Decimal

from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, JsonResponse
from django.template.loader import render_to_string
from django.contrib import messages
from django.utils import timezone
from django.views.decorators.http import require_POST
from django.views.decorators.csrf import csrf_exempt

from .models import (
    Producto,
    ConteoLinea,
    ConteoSesion,
    Proveedor,
    UnidadMedida,
)
from .forms import ProductoForm

from weasyprint import HTML
import openpyxl
from openpyxl.styles import Font


# =========================================================
# LISTA DE PRODUCTOS
# =========================================================

def lista_productos(request):
    productos = Producto.objects.all()
    return render(request, 'inventario/lista_productos.html', {'productos': productos})


# =========================================================
# CREAR PRODUCTO
# =========================================================

def crear_producto(request):
    if request.method == 'POST':
        form = ProductoForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('lista_productos')
    else:
        form = ProductoForm()
    return render(request, 'inventario/crear_producto.html', {'form': form})


# =========================================================
# EXPORTACIÓN XLSX DE PRODUCTOS
# =========================================================

def exportar_productos_xlsx(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Productos"

    encabezados = [
        "Código interno",
        "Nombre interno",
        "Nombre proveedor",
        "Categoría",
        "Unidad",
        "Proveedor",
        "IVA",
        "Precio compra",
        "Stock actual",
        "Stock mínimo",
        "Diferencia",
        "Activo",
    ]

    for col, titulo in enumerate(encabezados, 1):
        celda = ws.cell(row=1, column=col, value=titulo)
        celda.font = Font(bold=True)

    productos = Producto.objects.select_related(
        "categoria", "unidad_medida", "proveedor", "iva"
    ).all()

    fila = 2
    for p in productos:
        ws.cell(row=fila, column=1, value=p.codigo_interno)
        ws.cell(row=fila, column=2, value=p.nombre_interno)
        ws.cell(row=fila, column=3, value=p.nombre_proveedor)
        ws.cell(row=fila, column=4, value=p.categoria.nombre if p.categoria else "")
        ws.cell(row=fila, column=5, value=p.unidad_medida.abreviatura if p.unidad_medida else "")
        ws.cell(row=fila, column=6, value=p.proveedor.nombre if p.proveedor else "")
        ws.cell(row=fila, column=7, value=f"{p.iva.porcentaje}%" if p.iva else "")
        ws.cell(row=fila, column=8, value=float(p.precio_compra))
        ws.cell(row=fila, column=9, value=float(p.stock_actual))
        ws.cell(row=fila, column=10, value=float(p.stock_minimo))
        ws.cell(row=fila, column=11, value=float(p.stock_actual - p.stock_minimo))
        ws.cell(row=fila, column=12, value="Sí" if p.activo else "No")
        fila += 1

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="productos.xlsx"'
    wb.save(response)
    return response


# =========================================================
# EXPORTACIÓN PDF DE PRODUCTOS
# =========================================================

def exportar_productos_pdf(request):
    productos = Producto.objects.all()

    html_string = render_to_string(
        "inventario/productos_pdf.html",
        {"productos": productos}
    )

    pdf = HTML(string=html_string).write_pdf()

    response = HttpResponse(pdf, content_type="application/pdf")
    response["Content-Disposition"] = "inline; filename=productos.pdf"
    return response


# =========================================================
# CONTEO POR CÓDIGO DE BARRAS
# =========================================================

def conteo_codigo_barras(request, sesion_id):
    sesion = get_object_or_404(ConteoSesion, id=sesion_id)

    # AJAX: añadir línea
    if request.method == "POST":

        codigo = request.POST.get("codigo_barras")
        cantidad = float(request.POST.get("cantidad", 1))

        try:
            producto = Producto.objects.get(codigo_barras=codigo)
        except Producto.DoesNotExist:
            return JsonResponse({"error": "Código no encontrado"}, status=400)

        linea, creada = ConteoLinea.objects.get_or_create(
            sesion=sesion,
            producto=producto,
            defaults={
                "codigo_barras": codigo,
                "precio_coste": producto.precio_compra,
                "stock_teorico": producto.stock_actual,
                "stock_contado": 0,
            }
        )

        linea.stock_contado += cantidad
        linea.save()

        return JsonResponse({
            "producto": producto.nombre_interno,
            "stock_teorico": float(linea.stock_teorico),
            "stock_contado": float(linea.stock_contado),
            "diferencia": float(linea.diferencia),
            "precio_stock": float(linea.precio_coste),
            "importe": float(linea.importe_diferencia),
        })

    # Página normal
    lineas = sesion.lineas.all()

    total_teorico = sum(l.importe_teorico for l in lineas)
    total_contado = sum(l.importe_contado for l in lineas)
    total_diferencia = sum(l.importe_diferencia for l in lineas)

    total_aumento = sum(l.importe_diferencia for l in lineas if l.diferencia > 0)
    total_disminucion = sum(abs(l.importe_diferencia) for l in lineas if l.diferencia < 0)

    contexto = {
        "sesion": sesion,
        "lineas": lineas,
        "total_teorico": total_teorico,
        "total_contado": total_contado,
        "total_diferencia": total_diferencia,
        "total_aumento": total_aumento,
        "total_disminucion": total_disminucion,
    }

    return render(request, "inventario/conteo_codigo_barras.html", contexto)


# =========================================================
# AJAX: BUSCAR PRODUCTO POR CÓDIGO
# =========================================================

def buscar_producto_por_codigo(request):
    codigo = request.GET.get("codigo")

    if not codigo:
        return JsonResponse({"ok": False, "error": "Código vacío"})

    try:
        p = Producto.objects.get(codigo_barras=codigo)
    except Producto.DoesNotExist:
        return JsonResponse({"ok": False, "error": "Producto no encontrado"})

    return JsonResponse({
        "ok": True,
        "producto_id": p.id,
        "producto_nombre": p.nombre_interno,
        "precio_coste": float(p.precio_compra),
        "stock_teorico": float(p.stock_actual),
        "importe_teorico": float(p.precio_compra * p.stock_actual),
    })


# =========================================================
# AJAX: EDICIÓN RÁPIDA DEL STOCK CONTADO
# =========================================================

@csrf_exempt
def actualizar_stock_contado(request, linea_id):
    if request.method != "POST":
        return JsonResponse({"ok": False, "error": "Método no permitido"})

    try:
        linea = ConteoLinea.objects.get(id=linea_id)
    except ConteoLinea.DoesNotExist:
        return JsonResponse({"ok": False, "error": "Línea no encontrada"})

    if linea.sesion.estado == "cerrada":
        return JsonResponse({"ok": False, "error": "La sesión está cerrada"})

    try:
        nuevo_valor = Decimal(request.POST.get("valor"))
    except:
        return JsonResponse({"ok": False, "error": "Valor inválido"})

    linea.stock_contado = nuevo_valor
    linea.save()

    return JsonResponse({
        "ok": True,
        "importe_contado": float(linea.importe_contado),
        "diferencia": float(linea.diferencia),
        "importe_diferencia": float(linea.importe_diferencia),
    })


# =========================================================
# IMPORTACIÓN XLSX
# =========================================================

def importar_conteo_xlsx(request, sesion_id):
    sesion = get_object_or_404(ConteoSesion, id=sesion_id)

    if request.method != "POST" or "archivo" not in request.FILES:
        return HttpResponse("Debe enviar un archivo XLSX mediante POST.", status=400)

    archivo = request.FILES["archivo"]

    try:
        wb = openpyxl.load_workbook(archivo, data_only=True)
    except Exception:
        return HttpResponse("No se pudo leer el archivo XLSX.", status=400)

    ws = wb.active

    header_row = ws[1]
    headers = [str(c.value).strip() if c.value is not None else "" for c in header_row]

    def idx(nombre):
        for i, h in enumerate(headers):
            if h.lower() == nombre.lower():
                return i
        return None

    idx_codigo = idx("Codigo barras")
    idx_producto = idx("Producto")
    idx_precio = idx("Precio coste")
    idx_stock_teorico = idx("Stock teorico")
    idx_stock_contado = idx("Stock contado")
    idx_motivo = idx("Motivo")
    idx_ubicacion = idx("Ubicacion")
    idx_eliminar = idx("¿Eliminar?")

    if idx_codigo is None:
        return HttpResponse("El archivo XLSX debe contener la columna 'Codigo barras'.", status=400)

    for row in ws.iter_rows(min_row=2, values_only=True):

        codigo_barras = row[idx_codigo] if idx_codigo is not None else None
        if not codigo_barras:
            continue

        codigo_barras = str(codigo_barras).strip()

        nombre_producto = (
            str(row[idx_producto]).strip()
            if idx_producto is not None and row[idx_producto] is not None
            else f"Producto sin nombre ({codigo_barras})"
        )

        precio_coste = Decimal("0")
        if idx_precio is not None and row[idx_precio] is not None:
            try:
                precio_coste = Decimal(str(row[idx_precio]))
            except Exception:
                precio_coste = Decimal("0")

        stock_teorico = Decimal("0")
        if idx_stock_teorico is not None and row[idx_stock_teorico] is not None:
            try:
                stock_teorico = Decimal(str(row[idx_stock_teorico]))
            except Exception:
                stock_teorico = Decimal("0")

        stock_contado = None
        if idx_stock_contado is not None and row[idx_stock_contado] is not None:
            try:
                stock_contado = Decimal(str(row[idx_stock_contado]))
            except Exception:
                stock_contado = None

        motivo = (
            str(row[idx_motivo]).strip()
            if idx_motivo is not None and row[idx_motivo] is not None
            else ""
        )

        ubicacion = (
            str(row[idx_ubicacion]).strip()
            if idx_ubicacion is not None and row[idx_ubicacion] is not None
            else ""
        )

        eliminar_flag = (
            str(row[idx_eliminar]).strip().lower()
            if idx_eliminar is not None and row[idx_eliminar] is not None
            else ""
        )

        eliminar = eliminar_flag in ("si", "sí", "yes", "y", "1", "true")

        try:
            producto = Producto.objects.get(codigo_barras=codigo_barras)
        except Producto.DoesNotExist:

            proveedor_default = Proveedor.objects.first()
            unidad_default = UnidadMedida.objects.first()

            producto = Producto.objects.create(
                nombre_interno=nombre_producto,
                nombre_proveedor=nombre_producto,
                codigo_barras=codigo_barras,
                codigo_interno=codigo_barras,
                precio_compra=precio_coste,
                stock_actual=stock_teorico,
                activo=True,
                proveedor=proveedor_default,
                unidad_medida=unidad_default,
            )

        try:
            linea_existente = ConteoLinea.objects.get(sesion=sesion, producto=producto)
        except ConteoLinea.DoesNotExist:
            linea_existente = None

        if eliminar:
            if linea_existente is not None:
                linea_existente.delete()
            continue

        linea, creada = ConteoLinea.objects.get_or_create(
            sesion=sesion,
            producto=producto,
            defaults={
                "codigo_barras": codigo_barras,
                "stock_teorico": stock_teorico,
                "stock_contado": stock_contado if stock_contado is not None else Decimal("0"),
            },
        )

        linea.codigo_barras = codigo_barras
        linea.stock_teorico = stock_teorico

        if stock_contado is not None:
            linea.stock_contado = stock_contado
        elif linea.stock_contado is None:
            linea.stock_contado = Decimal("0")

        if hasattr(linea, "motivo"):
            linea.motivo = motivo
        if hasattr(linea, "ubicacion"):
            linea.ubicacion = ubicacion

        linea.save()

    return HttpResponse("Importación de conteo XLSX completada correctamente.")


# =========================================================
# FINALIZAR CONTEO
# =========================================================

@require_POST
def finalizar_conteo(request, sesion_id):
    sesion = get_object_or_404(ConteoSesion, id=sesion_id)

    if sesion.estado == "cerrada":
        messages.warning(request, "Esta sesión ya está cerrada.")
        return redirect("conteo_codigo_barras", sesion_id=sesion.id)

    sesion.generar_asientos_conteo(usuario=request.user)

    sesion.estado = "cerrada"
    sesion.fecha_cierre = timezone.now()
    sesion.usuario_cierre = request.user
    sesion.save()

    messages.success(
        request,
        "Conteo finalizado correctamente. Se han generado los asientos contables."
    )

    return redirect("conteo_codigo_barras", sesion_id=sesion.id)


# =========================================================
# REABRIR CONTEO
# =========================================================

def reabrir_conteo(request, sesion_id):
    sesion = get_object_or_404(ConteoSesion, id=sesion_id)

    if sesion.estado != "cerrada":
        messages.warning(request, "Solo se pueden reabrir sesiones cerradas.")
        return redirect("admin:inventario_app_conteosesion_changelist")

    for asiento in sesion.asientos.all():
        asiento.revertir(usuario=request.user)

    sesion.estado = "reabierta"
    sesion.fecha_reapertura = timezone.now()
    sesion.usuario_reapertura = request.user
    sesion.save()

    messages.success(
        request,
        "La sesión ha sido reabierta y los asientos contables han sido revertidos."
    )

    return redirect("conteo_codigo_barras", sesion_id=sesion.id)


# =========================================================
# EXPORTACIÓN XLSX DEL CONTEO
# =========================================================

def exportar_conteo_xlsx(request, sesion_id):
    sesion = get_object_or_404(ConteoSesion, id=sesion_id)
    lineas = sesion.lineas.all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Conteo {sesion.id}"

    encabezados = [
        "Código",
        "Producto",
        "Precio coste",
        "Stock teórico",
        "Importe teórico",
        "Stock contado",
        "Importe contado",
        "Diferencia",
        "Aumento €",
        "Disminución €",
    ]

    for col, titulo in enumerate(encabezados, 1):
        celda = ws.cell(row=1, column=col, value=titulo)
        celda.font = Font(bold=True)

    for l in lineas:
        ws.cell(row=fila, column=1, value=l.codigo_barras)
        ws.cell(row=fila, column=2, value=l.producto.nombre_interno)
        ws.cell(row=fila, column=3, value=float(l.precio_coste))
        ws.cell(row=fila, column=4, value=float(l.stock_teorico))
        ws.cell(row=fila, column=5, value=float(l.importe_teorico))
        ws.cell(row=fila, column=6, value=float(l.stock_contado))
        ws.cell(row=fila, column=7, value=float(l.importe_contado))
        ws.cell(row=fila, column=8, value=float(l.diferencia))
        ws.cell(row=fila, column=9, value=float(l.importe_diferencia if l.diferencia > 0 else 0))
        ws.cell(row=fila, column=10, value=float(abs(l.importe_diferencia) if l.diferencia < 0 else 0))
        fila += 1

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="conteo_{sesion.id}.xlsx"'
    wb.save(response)
    return response


# =========================================================
# EXPORTACIÓN PDF DEL CONTEO
# =========================================================

def exportar_conteo_pdf(request, sesion_id):
    sesion = get_object_or_404(ConteoSesion, id=sesion_id)
    lineas = sesion.lineas.all()

    html_string = render_to_string(
        "inventario/conteo_pdf.html",
        {
            "sesion": sesion,
            "lineas": lineas,
            "total_teorico": sesion.total_teorico,
            "total_contado": sesion.total_contado,
            "total_diferencia": sesion.total_diferencia,
            "total_aumento": sesion.total_aumento,
            "total_disminucion": sesion.total_disminucion,
        }
    )